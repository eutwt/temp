import openpyxl
import openpyxl.utils

def assert_ssn_column_is_character_first_sheet(excel_file, header_row=1):
    """
    Asserts that an Excel file contains a column named 'SSN' and all cells
    in that column (below the header) on the first sheet are formatted
    as character/text in Excel.

    Args:
        excel_file (str): The path to the Excel file.
        header_row (int): The row number where the column headers are located.
                          Defaults to 1 (the first row).

    Raises:
        FileNotFoundError: If the specified Excel file does not exist.
        IndexError: If the workbook is empty (no sheets).
        ValueError: If the 'SSN' column is not found in the header row on the first sheet.
        AssertionError: If any cell in the 'SSN' column on the first sheet
                        is not formatted as text in Excel.
    """
    ssn_column_index = -1
    ssn_column_letter = None
    sheet = None # Initialize sheet variable

    try:
        workbook = openpyxl.load_workbook(excel_file)

        # Get the first sheet
        if not workbook.sheetnames:
             raise IndexError(f"The workbook '{excel_file}' contains no sheets.")
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]

        print(f"Using the first sheet: '{sheet_name}'")

        # 1. Find the column index and letter of the 'SSN' column in the header row
        header_row_cells = sheet[header_row]
        for cell in header_row_cells:
            if isinstance(cell.value, str) and cell.value.strip().upper() == 'SSN':
                ssn_column_index = cell.column
                ssn_column_letter = cell.column_letter
                break

        if ssn_column_index == -1:
            raise ValueError(f"Column named 'SSN' not found in row {header_row} of the first sheet '{sheet_name}'.")

        print(f"Found 'SSN' column at column {ssn_column_letter} (index {ssn_column_index})")

        # 2. Iterate through cells in the SSN column (starting from the row after the header)
        for row_index in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=ssn_column_index)

            # Check if the number format is '@' (text) and the cell is not empty
            if cell.number_format != '@' and cell.value is not None:
                 raise AssertionError(
                    f"Cell {cell.coordinate} in the 'SSN' column on sheet '{sheet_name}' "
                    f"is not formatted as text in Excel. Number Format: {cell.number_format}, "
                    f"Value: {cell.value}"
                )
            # Optional: You might add a check here to ensure non-empty cells actually
            # contain string-like data if you have strict requirements beyond Excel's formatting.
            # For example:
            # if cell.value is not None and not isinstance(cell.value, str):
            #      raise AssertionError(f"Cell {cell.coordinate} in the 'SSN' column contains a non-string value: {cell.value}")


        print(f"Assertion successful: All cells in the 'SSN' column on the first sheet "
              f"'{sheet_name}' are formatted as character/text in Excel.")

    except FileNotFoundError:
        print(f"Error: File not found at {excel_file}")
        raise # Re-raise the exception after printing
    except IndexError:
        print(f"Error: The workbook '{excel_file}' is empty and contains no sheets.")
        raise # Re-raise the exception after printing
    except ValueError as ve:
        print(f"Error: {ve}")
        raise # Re-raise the exception after printing
    except AssertionError as ae:
        print(f"Assertion Failed: {ae}")
        raise # Re-raise the exception after printing
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        raise # Re-raise any other unexpected exceptions

# Example Usage:
excel_file_path = 'your_excel_file.xlsx' # Replace with the path to your file

try:
    assert_ssn_column_is_character_first_sheet(excel_file_path)
    print("Validation passed.")
except (FileNotFoundError, IndexError, ValueError, AssertionError) as e:
    print("Validation failed.")
    # The specific error message is printed within the function
